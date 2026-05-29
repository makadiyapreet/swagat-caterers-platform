"""
Section 12: Excel & CSV Export Functions
Generates styled Excel exports for events, members, and bookings.
"""
from django.http import HttpResponse
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .decorators import require_role
from .models import CateringEvent, Member, MemberLog, Booking
from django.utils import timezone
from datetime import datetime
import csv


def _style_header(ws, headers):
    """Apply brand styling to Excel header row."""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        gold_fill = PatternFill(start_color='1A1A1A', end_color='1A1A1A', fill_type='solid')
        gold_font = Font(name='Calibri', bold=True, color='C9A84C', size=11)
        thin_border = Border(
            left=Side(style='thin', color='333333'),
            right=Side(style='thin', color='333333'),
            top=Side(style='thin', color='333333'),
            bottom=Side(style='thin', color='333333'),
        )

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = gold_font
            cell.fill = gold_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = max(len(header) + 5, 15)

    except ImportError:
        # openpyxl not installed — just write plain headers
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)


@login_required
@require_role(['admin', 'manager'])
def export_events_excel(request):
    """Export events to styled Excel file."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse("openpyxl not installed. Run: pip install openpyxl", status=500)

    # Get date range from query params
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    events = CateringEvent.objects.all().order_by('-date')
    
    if start_date:
        events = events.filter(date__gte=start_date)
    if end_date:
        events = events.filter(date__lte=end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Events'

    headers = ['ID', 'Title', 'Date', 'Venue', 'Guests', 'Event Type', 
               'Status', 'Rate/Plate', 'Advance', 'Total Cost', 'Pending', 'Contact']
    _style_header(ws, headers)

    for row_num, event in enumerate(events, 2):
        ws.cell(row=row_num, column=1, value=event.id)
        ws.cell(row=row_num, column=2, value=event.title)
        ws.cell(row=row_num, column=3, value=str(event.date))
        ws.cell(row=row_num, column=4, value=event.venue)
        ws.cell(row=row_num, column=5, value=event.guests)
        ws.cell(row=row_num, column=6, value=event.event_type)
        ws.cell(row=row_num, column=7, value=event.status)
        ws.cell(row=row_num, column=8, value=float(event.rate))
        ws.cell(row=row_num, column=9, value=float(event.advance_amount))
        ws.cell(row=row_num, column=10, value=float(event.total_cost))
        ws.cell(row=row_num, column=11, value=float(event.pending_amount))
        ws.cell(row=row_num, column=12, value=event.contact_number or '')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'swagat_events_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@require_role(['admin', 'manager'])
def export_members_excel(request):
    """Export members to styled Excel file."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse("openpyxl not installed", status=500)

    members = Member.objects.all().order_by('name')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Staff Members'

    headers = ['ID', 'Name', 'Phone', 'Daily Wage', 'Created At']
    _style_header(ws, headers)

    for row_num, member in enumerate(members, 2):
        ws.cell(row=row_num, column=1, value=member.id)
        ws.cell(row=row_num, column=2, value=member.name)
        ws.cell(row=row_num, column=3, value=member.phone)
        ws.cell(row=row_num, column=4, value=float(member.daily_wage))
        ws.cell(row=row_num, column=5, value=str(member.created_at) if hasattr(member, 'created_at') else '')

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'swagat_members_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@require_role(['admin', 'manager'])
def export_bookings_excel(request):
    """Export bookings to styled Excel file."""
    try:
        from openpyxl import Workbook
    except ImportError:
        return HttpResponse("openpyxl not installed", status=500)

    start_date = request.GET.get('start')
    end_date = request.GET.get('end')

    bookings = Booking.objects.all().order_by('-created_at')

    if start_date:
        bookings = bookings.filter(event_date__gte=start_date)
    if end_date:
        bookings = bookings.filter(event_date__lte=end_date)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Bookings'

    headers = ['ID', 'Name', 'Phone', 'Event Date', 'Event Type', 
               'Guests', 'Meal Time', 'Package', 'Venue', 'Message', 'Created At']
    _style_header(ws, headers)

    for row_num, b in enumerate(bookings, 2):
        ws.cell(row=row_num, column=1, value=b.id)
        ws.cell(row=row_num, column=2, value=b.name)
        ws.cell(row=row_num, column=3, value=b.phone)
        ws.cell(row=row_num, column=4, value=str(b.event_date))
        ws.cell(row=row_num, column=5, value=b.event_type)
        ws.cell(row=row_num, column=6, value=b.guest_count)
        ws.cell(row=row_num, column=7, value=b.meal_time or '')
        ws.cell(row=row_num, column=8, value=b.package_type)
        ws.cell(row=row_num, column=9, value=b.venue or '')
        ws.cell(row=row_num, column=10, value=b.message or '')
        ws.cell(row=row_num, column=11, value=str(b.created_at))

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'swagat_bookings_{timezone.now().strftime("%Y%m%d")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
@require_role(['admin', 'manager'])
def export_events_csv(request):
    """Export events as CSV (simpler alternative)."""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="events_{timezone.now().strftime("%Y%m%d")}.csv"'

    writer = csv.writer(response)
    writer.writerow(['ID', 'Title', 'Date', 'Venue', 'Guests', 'Event Type', 
                     'Status', 'Rate', 'Advance', 'Total', 'Pending', 'Contact'])

    for event in CateringEvent.objects.all().order_by('-date'):
        writer.writerow([
            event.id, event.title, event.date, event.venue, event.guests,
            event.event_type, event.status, event.rate, event.advance_amount,
            event.total_cost, event.pending_amount, event.contact_number or ''
        ])

    return response


# --- Export Preview Pages ---

@login_required
@require_role(['admin', 'manager'])
def export_events_preview(request):
    """Preview page for events export."""
    return render(request, "export_events_preview.html")


@login_required
@require_role(['admin', 'manager'])
def export_bookings_preview(request):
    """Preview page for bookings export."""
    return render(request, "export_bookings_preview.html")


# --- Export Preview API (JSON) ---

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response as DRFResponse
from dateutil.relativedelta import relativedelta


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_events_api(request):
    """Return events data as JSON for preview. Auto-excludes data older than 6 months."""
    user_type = getattr(request.user, 'user_type', 'customer')
    if user_type not in ('admin', 'manager'):
        return DRFResponse({'error': 'Permission denied'}, status=403)
    
    six_months_ago = timezone.now().date() - relativedelta(months=6)
    events = CateringEvent.objects.filter(date__gte=six_months_ago).order_by('-date')
    
    data = [{
        'id': e.id,
        'title': e.title,
        'date': str(e.date),
        'venue': e.venue,
        'guests': e.guests,
        'event_type': e.event_type,
        'status': e.status,
        'rate': float(e.rate),
        'advance': float(e.advance_amount),
        'total': float(e.total_cost),
        'pending': float(e.pending_amount),
        'contact': e.contact_number or '',
    } for e in events]
    
    return DRFResponse(data)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_bookings_api(request):
    """Return bookings data as JSON for preview. Auto-excludes data older than 6 months."""
    user_type = getattr(request.user, 'user_type', 'customer')
    if user_type not in ('admin', 'manager'):
        return DRFResponse({'error': 'Permission denied'}, status=403)
    
    six_months_ago = timezone.now().date() - relativedelta(months=6)
    bookings = Booking.objects.filter(event_date__gte=six_months_ago).order_by('-created_at')
    
    data = [{
        'id': b.id,
        'name': b.name,
        'phone': b.phone,
        'event_date': str(b.event_date),
        'event_type': b.event_type,
        'guests': b.guest_count,
        'meal_time': b.meal_time or '',
        'package': b.package_type,
        'venue': b.venue or '',
        'message': b.message or '',
        'created_at': str(b.created_at),
    } for b in bookings]
    
    return DRFResponse(data)

